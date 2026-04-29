"""
Excel Form Submit Loop — helper module for WindAutomateX.

Loads rows from an Excel (.xlsx) or CSV file and submits each row into a
target form using pywinauto / pyautogui.

Public API
----------
run_excel_form_loop(config, engine) -> dict
    Entry-point called by the main execution engine.

get_excel_headers(file_path, sheet_name) -> list[str]
    Return the header row from an Excel or CSV file.
"""

import csv
import json
import logging
import os
import time

logger = logging.getLogger(__name__)

# Offset used when converting 1-based sheet row numbers to 0-based list indices
# that also account for a header row: header(1) + 1-based(1) = offset of 2.
_HEADER_ROW_OFFSET = 2

# ---------------------------------------------------------------------------
# Resume-state file helpers
# ---------------------------------------------------------------------------

def _state_path(file_path: str) -> str:
    """Return the path of the resume-state JSON file for *file_path*."""
    base = os.path.splitext(file_path)[0]
    return base + "_wax_resume.json"


def _load_resume_state(file_path: str) -> int:
    """Return the last successfully processed row index (0-based), or -1."""
    state_file = _state_path(file_path)
    try:
        with open(state_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            return int(data.get("last_row", -1))
    except (FileNotFoundError, json.JSONDecodeError, ValueError):
        return -1


def _save_resume_state(file_path: str, row_index: int) -> None:
    """Persist the last successfully processed row index (0-based)."""
    state_file = _state_path(file_path)
    try:
        with open(state_file, "w", encoding="utf-8") as f:
            json.dump({"last_row": row_index}, f)
    except OSError as e:
        logger.warning(f"Could not save resume state: {e}")


def _clear_resume_state(file_path: str) -> None:
    state_file = _state_path(file_path)
    try:
        os.remove(state_file)
    except FileNotFoundError:
        pass


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_excel_rows(
    file_path: str,
    sheet_name: str = "Sheet1",
    has_header: bool = True,
    start_row: int = 2,
    end_row: int | None = None,
) -> list[dict]:
    """
    Load rows from an Excel (.xlsx / .xls) or CSV file.

    Returns a list of dicts mapping header names (or column indices) to cell
    values.  *start_row* and *end_row* are 1-based, inclusive.
    """
    if not file_path:
        raise ValueError("filePath is required")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        return _load_xlsx_rows(file_path, sheet_name, has_header, start_row, end_row)
    elif ext == ".csv":
        return _load_csv_rows(file_path, has_header, start_row, end_row)
    else:
        raise ValueError(f"Unsupported file extension: {ext}")


def _load_xlsx_rows(
    file_path: str,
    sheet_name: str,
    has_header: bool,
    start_row: int,
    end_row: int | None,
) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for xlsx support. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {available}"
        )

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return []

    headers: list[str]
    data_rows: list[tuple]

    if has_header:
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(all_rows[0])]
        data_rows = all_rows[1:]
        # start_row is 1-based counting from the first DATA row
        slice_start = max(0, start_row - _HEADER_ROW_OFFSET)
    else:
        headers = [f"col_{i}" for i in range(len(all_rows[0]))]
        data_rows = all_rows
        slice_start = max(0, start_row - 1)

    slice_end = (end_row - 1) if end_row is not None else None
    if slice_end is not None:
        if has_header:
            slice_end = slice_end - 1  # convert sheet row → data_rows index
        data_rows = data_rows[slice_start:slice_end]
    else:
        data_rows = data_rows[slice_start:]

    return [
        {headers[i]: (str(cell) if cell is not None else "") for i, cell in enumerate(row)}
        for row in data_rows
    ]


def _load_csv_rows(
    file_path: str,
    has_header: bool,
    start_row: int,
    end_row: int | None,
) -> list[dict]:
    rows: list[dict] = []
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        raw_rows = list(reader)

    if not raw_rows:
        return []

    headers: list[str]
    data_rows: list[list[str]]

    if has_header:
        headers = raw_rows[0]
        data_rows = raw_rows[1:]
        slice_start = max(0, start_row - _HEADER_ROW_OFFSET)
    else:
        headers = [f"col_{i}" for i in range(len(raw_rows[0]))]
        data_rows = raw_rows
        slice_start = max(0, start_row - 1)

    slice_end: int | None = None
    if end_row is not None:
        slice_end = (end_row - 1) - (1 if has_header else 0)

    if slice_end is not None:
        data_rows = data_rows[slice_start:slice_end]
    else:
        data_rows = data_rows[slice_start:]

    for row in data_rows:
        rows.append({headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))})
    return rows


# ---------------------------------------------------------------------------
# Header inspection
# ---------------------------------------------------------------------------

def get_excel_headers(file_path: str, sheet_name: str | None = None) -> list[str]:
    """
    Return the list of column header names from the first row of *file_path*.

    Works for ``.xlsx`` / ``.xls`` (via openpyxl) and ``.csv`` files.
    If *sheet_name* is provided and exists the matching sheet is used;
    otherwise the workbook's active (first) sheet is used.
    """
    if not file_path:
        raise ValueError("file_path is required")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        with open(file_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            first_row = next(reader, [])
        return [str(h).strip() for h in first_row if str(h).strip()]

    # XLSX / XLS
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for xlsx support. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    wb.close()
    return [str(h).strip() for h in first_row if h is not None and str(h).strip()]


# ---------------------------------------------------------------------------
# Form interaction helpers
# ---------------------------------------------------------------------------

def fill_field(selector: str, value: str, input_type: str, engine) -> None:
    """
    Fill a single form field identified by *selector* with *value*.

    Uses pywinauto (element title / auto_id) for native controls, or
    pyautogui tab-typing as a fallback.

    When *input_type* is ``"coordinate"``, *selector* must be ``"x,y"``
    (absolute screen coordinates).  The helper clicks that position and
    then types the value.
    """
    if input_type == "coordinate":
        parts = selector.split(",")
        if len(parts) != 2:
            raise ValueError(
                f"coordinate selector must be 'x,y' with numeric values, got: {selector!r}"
            )
        try:
            x, y = int(parts[0].strip()), int(parts[1].strip())
        except ValueError:
            raise ValueError(
                f"coordinate selector 'x,y' must contain integers, got: {selector!r}"
            )
        if engine.pyautogui_available:
            import pyautogui
            pyautogui.click(x, y)
            time.sleep(0.1)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.write(str(value), interval=0.03)
        return
    if input_type == "text":
        if engine.pywinauto_available:
            from pywinauto import Application, findwindows
            wins = findwindows.find_windows(active_only=True)
            if wins:
                app = Application(backend="uia").connect(handle=wins[0])
                win = app.top_window()
                try:
                    ctrl = win.child_window(auto_id=selector)
                    ctrl.set_edit_text(value)
                    return
                except Exception:
                    pass
                try:
                    ctrl = win.child_window(title=selector, control_type="Edit")
                    ctrl.set_edit_text(value)
                    return
                except Exception:
                    pass
        # Fallback: use pyautogui to type into whichever control currently has focus
        if engine.pyautogui_available:
            import pyautogui
            pyautogui.hotkey("ctrl", "a")
            pyautogui.write(str(value), interval=0.03)
    elif input_type == "dropdown":
        if engine.pywinauto_available:
            from pywinauto import Application, findwindows
            wins = findwindows.find_windows(active_only=True)
            if wins:
                app = Application(backend="uia").connect(handle=wins[0])
                win = app.top_window()
                try:
                    combo = win.child_window(auto_id=selector, control_type="ComboBox")
                    combo.select(value)
                    return
                except Exception:
                    pass
    elif input_type == "checkbox":
        if engine.pywinauto_available:
            from pywinauto import Application, findwindows
            wins = findwindows.find_windows(active_only=True)
            if wins:
                app = Application(backend="uia").connect(handle=wins[0])
                win = app.top_window()
                try:
                    chk = win.child_window(auto_id=selector, control_type="CheckBox")
                    desired = str(value).lower() in ("1", "true", "yes")
                    current = chk.get_toggle_state() == 1
                    if current != desired:
                        chk.toggle()
                    return
                except Exception:
                    pass


def click_submit(selector: str, engine) -> None:
    """Click the submit button identified by *selector*.

    *selector* may be an element auto_id / title string, or an ``"x,y"``
    coordinate string produced by the screen coordinate picker.
    """
    # Coordinate shortcut: "x,y" format
    parts = selector.split(",")
    if len(parts) == 2:
        try:
            x, y = int(parts[0].strip()), int(parts[1].strip())
            if engine.pyautogui_available:
                import pyautogui
                pyautogui.click(x, y)
            return
        except ValueError:
            pass  # not a coordinate selector — fall through to element-based logic
    if engine.pywinauto_available:
        from pywinauto import Application, findwindows
        wins = findwindows.find_windows(active_only=True)
        if wins:
            app = Application(backend="uia").connect(handle=wins[0])
            win = app.top_window()
            try:
                btn = win.child_window(auto_id=selector)
                btn.click_input()
                return
            except Exception:
                pass
            try:
                btn = win.child_window(title=selector, control_type="Button")
                btn.click_input()
                return
            except Exception:
                pass
    if engine.pyautogui_available:
        import pyautogui
        pyautogui.press("enter")


def _press_keyboard_shortcut(shortcut: str, engine) -> None:
    """Press a keyboard shortcut using pyautogui.

    *shortcut* should be a ``+``-separated string of key names such as
    ``"enter"``, ``"ctrl+s"``, or ``"alt+f4"``.
    """
    if not shortcut:
        return
    if engine.pyautogui_available:
        import pyautogui
        keys = [k.strip().lower() for k in shortcut.split("+") if k.strip()]
        if not keys:
            return
        if len(keys) == 1:
            pyautogui.press(keys[0])
        else:
            pyautogui.hotkey(*keys)


def _wait_for_success_text(success_text: str, timeout_ms: int, engine) -> bool:
    """Poll the active window for *success_text* until *timeout_ms* elapses."""
    if not success_text or not engine.pywinauto_available:
        return True
    from pywinauto import Application, findwindows
    deadline = time.monotonic() + timeout_ms / 1000.0
    while time.monotonic() < deadline:
        try:
            wins = findwindows.find_windows(active_only=True)
            if wins:
                app = Application(backend="uia").connect(handle=wins[0])
                win = app.top_window()
                if success_text.lower() in win.window_text().lower():
                    return True
        except Exception:
            pass
        time.sleep(0.3)
    return False


def _take_screenshot(row_index: int, file_path: str, engine) -> None:
    """Save a failure screenshot next to the source file."""
    if not engine.pyautogui_available:
        return
    import pyautogui
    base_dir = os.path.dirname(file_path) or "."
    shot_path = os.path.join(base_dir, f"wax_failure_row{row_index + 1}_{int(time.time())}.png")
    try:
        pyautogui.screenshot(shot_path)
        logger.info(f"Screenshot saved: {shot_path}")
    except Exception as e:
        logger.warning(f"Screenshot failed: {e}")


# ---------------------------------------------------------------------------
# Per-row processing
# ---------------------------------------------------------------------------

def process_row(
    row: dict,
    row_index: int,
    config: dict,
    engine,
) -> dict:
    """
    Fill fields, click submit, wait for success for a single *row*.

    Returns {"success": bool, "message": str}.
    """
    mappings: list[dict] = config.get("mappings", [])
    submit_actions: list[dict] = config.get("submitActions") or []
    submit_selector: str = config.get("submitSelector", "")  # backward compat
    wait_after_submit: int = int(config.get("waitAfterSubmit", 1500))
    success_text: str = config.get("successText", "") or ""
    clear_form: bool = bool(config.get("clearFormBeforeNextRow", False))

    try:
        # Optionally clear the form first
        if clear_form and engine.pywinauto_available:
            from pywinauto import Application, findwindows
            wins = findwindows.find_windows(active_only=True)
            if wins:
                app = Application(backend="uia").connect(handle=wins[0])
                win = app.top_window()
                for edit in win.children(control_type="Edit"):
                    try:
                        edit.set_edit_text("")
                    except Exception:
                        pass

        # Fill each mapped field
        for mapping in mappings:
            column: str = mapping.get("column", "")
            selector: str = mapping.get("selector", "")
            input_type: str = mapping.get("inputType", "text")
            if not column or not selector:
                continue
            value = row.get(column, "")
            fill_field(selector, str(value), input_type, engine)

        # Execute submit actions (new multi-action list takes precedence)
        if submit_actions:
            for action in submit_actions:
                action_type = action.get("type", "coordinate")
                value = action.get("value", "")
                if not value and action_type != "delay":
                    logger.warning(f"Row {row_index + 1}: submit action of type '{action_type}' has no value, skipping")
                    continue
                if action_type == "coordinate":
                    click_submit(value, engine)
                elif action_type == "keyboard":
                    _press_keyboard_shortcut(value, engine)
                elif action_type == "type_text":
                    # value is the Excel column name; type its text into the active field
                    text_to_type = str(row.get(value, ""))
                    if engine.pyautogui_available:
                        import pyautogui
                        pyautogui.write(text_to_type, interval=0.03)
                elif action_type == "delay":
                    delay_ms = int(value) if value else 0
                    if delay_ms > 0:
                        import time
                        time.sleep(delay_ms / 1000.0)
                elif action_type == "tick_vr":
                    # value is the Excel column name containing comma-separated VR numbers
                    from vr_checkbox_ticker import tick_checkboxes_by_vr
                    vr_list_str = str(row.get(value, "")) if value else ""
                    # Resolve Item Code from its own Excel column (if configured)
                    item_code_column = str(action.get("itemCodeColumn", "")).strip()
                    item_code = str(row.get(item_code_column, "")).strip() if item_code_column else ""
                    if vr_list_str:
                        tick_result = tick_checkboxes_by_vr(
                            vr_list_str,
                            window_title=str(action.get("windowTitle", "")),
                            grid_roi=str(action.get("gridRoi", "")),
                            scroll_x=int(action.get("scrollX", 0)),
                            scroll_y=int(action.get("scrollY", 0)),
                            max_scroll_attempts=int(action.get("maxScrollAttempts", 20)),
                            scroll_step=int(action.get("scrollStep", 3)),
                            checkbox_offset=int(action.get("checkboxOffset", 40)),
                            item_code=item_code,
                            engine=engine,
                        )
                        if not tick_result.get("success", False):
                            logger.warning(
                                "Row %d: tick_vr action partial result: %s",
                                row_index + 1,
                                tick_result.get("message", ""),
                            )
                    else:
                        logger.warning(
                            "Row %d: tick_vr action: column '%s' is empty or not found",
                            row_index + 1,
                            value,
                        )
        elif submit_selector:
            # Backward-compatible: old single submitSelector
            click_submit(submit_selector, engine)
        else:
            return {"success": False, "message": "No submit actions configured"}

        # Wait after submit
        if success_text:
            found = _wait_for_success_text(success_text, wait_after_submit, engine)
            if not found:
                logger.warning(f"Row {row_index + 1}: success text '{success_text}' not found within {wait_after_submit}ms")
        else:
            time.sleep(wait_after_submit / 1000.0)

        return {"success": True, "message": f"Row {row_index + 1} submitted"}

    except Exception as e:
        return {"success": False, "message": str(e)}


# ---------------------------------------------------------------------------
# Main entry-point
# ---------------------------------------------------------------------------

def run_excel_form_loop_for_row(config: dict, row: dict, row_index: int, engine) -> dict:
    """
    Execute the Excel Form Submit Loop for a single pre-loaded *row*.

    Called by the executor when running in outer-loop mode — i.e. when every
    task step (not just the form-filling step) repeats once per Excel row.
    Unlike :func:`run_excel_form_loop`, this variant does **not** load or
    iterate over the Excel file; it receives the already-loaded row dict and
    its 0-based *row_index*.

    At least one submit action must still be configured; field *mappings* are
    optional (they may be empty when the calling task handles input via other
    step types).

    Parameters
    ----------
    config : dict
        Parsed ``config_json`` from the ``excel_form_submit_loop`` task step.
    row : dict
        Column-name → cell-value mapping for the current Excel row.
    row_index : int
        0-based index of *row* within the loaded data rows (used for logging
        and screenshot file-names).
    engine : WindAutomateXEngine
        Shared engine instance.

    Returns
    -------
    dict
        ``{"success": bool, "message": str}``
    """
    submit_actions = config.get("submitActions") or []
    if not submit_actions and not config.get("submitSelector"):
        return {"success": False, "message": "Either submitActions or submitSelector must be configured"}

    retry_count: int = int(config.get("retryCount", 2))
    save_screenshot: bool = bool(config.get("saveScreenshotOnFailure", False))
    file_path: str = config.get("filePath", "")

    result = None
    for attempt in range(retry_count + 1):
        result = process_row(row, row_index, config, engine)
        if result["success"]:
            break
        if attempt < retry_count:
            logger.warning(
                f"Row {row_index + 1} attempt {attempt + 1} failed: {result['message']} — retrying…"
            )
            time.sleep(0.5)

    if result and not result["success"] and save_screenshot and file_path:
        _take_screenshot(row_index, file_path, engine)

    return result if result else {"success": False, "message": "Unknown error"}


def run_excel_form_loop(config: dict, engine) -> dict:
    """
    Execute the full Excel Form Submit Loop step.

    Parameters
    ----------
    config : dict
        Parsed step configuration JSON.
    engine : WindAutomateXEngine
        The shared engine instance (provides pywinauto/pyautogui availability flags).

    Returns
    -------
    dict
        {"success": bool, "message": str}
    """
    file_path: str = config.get("filePath", "")
    sheet_name: str = config.get("sheetName", "Sheet1")
    has_header: bool = bool(config.get("hasHeader", True))
    start_row: int = int(config.get("startRow", 2))
    end_row_raw = config.get("endRow")
    end_row: int | None = int(end_row_raw) if end_row_raw is not None else None
    continue_on_error: bool = bool(config.get("continueOnError", True))
    retry_count: int = int(config.get("retryCount", 2))
    delay_between_rows: int = int(config.get("delayBetweenRows", 1000))
    save_screenshot: bool = bool(config.get("saveScreenshotOnFailure", False))
    resume: bool = bool(config.get("resumeFromLastRow", False))

    # Validate required fields
    if not file_path:
        return {"success": False, "message": "filePath is required"}
    submit_actions = config.get("submitActions") or []
    if not submit_actions and not config.get("submitSelector"):
        return {"success": False, "message": "At least one submit action is required"}

    # Load rows
    try:
        rows = load_excel_rows(file_path, sheet_name, has_header, start_row, end_row)
    except Exception as e:
        return {"success": False, "message": f"Failed to load file: {e}"}

    if not rows:
        return {"success": False, "message": "No data rows found in the specified range"}

    total = len(rows)
    print(json.dumps({"event": "excel_loop_start", "total_rows": total, "file": os.path.basename(file_path)}), flush=True)
    logger.info(f"Loaded {total} rows from {os.path.basename(file_path)}")

    # Resume support
    resume_from = 0
    if resume:
        last = _load_resume_state(file_path)
        if last >= 0 and last < total:
            resume_from = last + 1
            logger.info(f"Resuming from row {resume_from + 1} (last successful: {last + 1})")
            print(json.dumps({"event": "excel_loop_resume", "from_row": resume_from + 1}), flush=True)

    success_count = 0
    fail_count = 0

    for i in range(resume_from, total):
        row = rows[i]

        result = None
        for attempt in range(retry_count + 1):
            result = process_row(row, i, config, engine)
            if result["success"]:
                break
            if attempt < retry_count:
                logger.warning(f"Row {i + 1} attempt {attempt + 1} failed: {result['message']} — retrying…")
                time.sleep(0.5)

        if result and result["success"]:
            success_count += 1
            logger.info(f"Row {i + 1} success")
            print(json.dumps({"event": "excel_row_done", "row": i + 1, "success": True}), flush=True)
            if resume:
                _save_resume_state(file_path, i)
        else:
            fail_count += 1
            msg = result["message"] if result else "Unknown error"
            logger.error(f"Row {i + 1} failed: {msg}")
            print(json.dumps({"event": "excel_row_done", "row": i + 1, "success": False, "message": msg}), flush=True)
            if save_screenshot:
                _take_screenshot(i, file_path, engine)
            if not continue_on_error:
                return {
                    "success": False,
                    "message": f"Stopped at row {i + 1}: {msg}. Completed {success_count}/{total} rows.",
                }

        # Delay between rows (skip after the last row)
        if i < total - 1 and delay_between_rows > 0:
            time.sleep(delay_between_rows / 1000.0)

    # Clear resume state on full completion
    if resume:
        _clear_resume_state(file_path)

    summary = f"Completed {success_count}/{total} rows"
    if fail_count:
        summary += f" ({fail_count} failed)"
    logger.info(summary)
    print(json.dumps({"event": "excel_loop_done", "success": success_count, "failed": fail_count, "total": total}), flush=True)

    # Treat partial success as overall success when continueOnError is True
    overall_success = fail_count == 0 or continue_on_error
    return {"success": overall_success, "message": summary}
